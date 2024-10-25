import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import logging
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from tkinter.ttk import Progressbar

# Configuração do log
logging.basicConfig(filename='log_script.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Função para verificar o arquivo ODS
def verificar_arquivo_ods():
    caminho_ods = os.path.join(os.getcwd(), 'total.ods')
    if not os.path.exists(caminho_ods):
        logging.error("Arquivo total.ods não foi encontrado.")
        return None
    logging.info(f"Arquivo ODS encontrado: {caminho_ods}")
    return caminho_ods

# Função para carregar o arquivo modelo.xlsx
def carregar_modelo_excel():
    modelo_path = 'modelo.xlsx'
    if not os.path.exists(modelo_path):
        logging.error("O arquivo modelo.xlsx não foi encontrado.")
        return None
    logging.info(f"Arquivo modelo.xlsx encontrado: {modelo_path}")
    return modelo_path

# Função para carregar o arquivo area.txt
def carregar_area():
    area_path = 'area.txt'
    if not os.path.exists(area_path):
        logging.error("O arquivo area.txt não foi encontrado.")
        return None
    with open(area_path, 'r') as f:
        area = f.read().strip()
    logging.info(f"Área carregada: {area}")
    return area

# Função para carregar a lista de nomes e remover linhas vazias
def carregar_lista_nomes():
    lista_nomes_path = 'lista_nomes.txt'
    if not os.path.exists(lista_nomes_path):
        logging.error("O arquivo lista_nomes.txt não foi encontrado.")
        return None
    with open(lista_nomes_path, 'r') as f:
        nomes = [nome.strip() for nome in f.readlines() if nome.strip()]
    logging.info(f"Lista de nomes carregada: {len(nomes)} nomes encontrados.")
    return nomes

# Função para identificar duplicidades de primeiro nome
def identificar_duplicidade_nomes(lista_nomes):
    primeiro_nome_count = {}
    for nome in lista_nomes:
        primeiro_nome = nome.split()[0]
        if primeiro_nome in primeiro_nome_count:
            primeiro_nome_count[primeiro_nome] += 1
        else:
            primeiro_nome_count[primeiro_nome] = 1
    return primeiro_nome_count

# Função para buscar o nome completo para a célula B2
def obter_nome_completo(nome_lista, lista_nomes):
    for nome_completo in lista_nomes:
        if nome_completo.startswith(nome_lista):
            return nome_completo
    return nome_lista  # Retorna o nome sem alterações se não encontrar

# Função para criar abas no Excel com controle de duplicidade e nome completo na célula B2
def criar_abas_excel(progresso):
    caminho_arquivo = verificar_arquivo_ods()
    if not caminho_arquivo:
        messagebox.showwarning("Aviso", "Não foi encontrado o arquivo ODS.")
        return

    modelo_path = carregar_modelo_excel()
    if not modelo_path:
        messagebox.showerror("Erro", "O arquivo modelo.xlsx não foi encontrado.")
        return

    area = carregar_area()
    if not area:
        messagebox.showerror("Erro", "O arquivo area.txt não foi encontrado.")
        return

    lista_nomes = carregar_lista_nomes()
    if not lista_nomes:
        messagebox.showerror("Erro", "O arquivo lista_nomes.txt não foi encontrado.")
        return

    try:
        # Carregar o ODS e filtrar apenas os nomes que estão na lista
        df = pd.read_excel(caminho_arquivo, engine="odf")

        # Remover "(Titular)" dos nomes
        df['Nome completo'] = df['Nome completo'].str.replace(r"\(Titular\)", "", regex=True).str.strip()

        # Criar o modelo Excel
        wb = load_workbook(modelo_path)
        if 'primeiroNome' not in wb.sheetnames:
            messagebox.showerror("Erro", "A aba 'primeiroNome' não foi encontrada no modelo.")
            return

        aba_modelo = wb['primeiroNome']
        abas_existentes = {}

        # Identificar duplicidades de primeiro nome
        duplicidade_nomes = identificar_duplicidade_nomes(lista_nomes)
        progresso['maximum'] = len(lista_nomes)

        # Iterar sobre a lista de nomes
        for i, nome_lista in enumerate(lista_nomes, start=1):
            df_filtrado = df[df['Nome completo'].str.contains(nome_lista, na=False, case=False)]

            nome_lista_split = nome_lista.split()
            primeiro_nome = nome_lista_split[0]
            segundo_nome = nome_lista_split[1] if len(nome_lista_split) > 1 else ""

            if duplicidade_nomes[primeiro_nome] > 1 and segundo_nome:
                nome_aba = f"{primeiro_nome}-{segundo_nome}"
            else:
                nome_aba = primeiro_nome

            if nome_aba in abas_existentes:
                nova_aba = abas_existentes[nome_aba]
                ultima_linha = nova_aba.max_row + 1
            else:
                nova_aba = wb.copy_worksheet(aba_modelo)
                nova_aba.title = nome_aba
                abas_existentes[nome_aba] = nova_aba
                ultima_linha = 9
                nova_aba['B1'] = area

                # Adicionar o nome completo na célula B2
                nome_completo = obter_nome_completo(nome_lista, lista_nomes)
                nova_aba['B2'] = nome_completo

                # Substituir placeholders no modelo
                for linha in nova_aba.iter_rows():
                    for cell in linha:
                        if isinstance(cell.value, str):
                            if "<Nome Completo>" in cell.value:
                                cell.value = cell.value.replace("<Nome Completo>", nome_completo)

            if df_filtrado.empty:
                logging.info(f"Nome {nome_lista} não encontrado no ODS. Criando aba vazia.")

                # Inserir Totais na linha 11
                bold_font = Font(bold=True, name="Tahoma", size=10)
                nova_aba["F11"].value = "Total"
                nova_aba["G11"].value = "0,00%"
                nova_aba["H11"].value = "0,00%"

                # Aplicar fonte em negrito
                nova_aba["F11"].font = bold_font
                nova_aba["G11"].font = bold_font
                nova_aba["H11"].font = bold_font

                logging.info("Linha de Total adicionada na linha 11 para caso de aba vazia.")
            else:
                colunas = ['Nome', 'Andamento', 'Início', 'Término', 'Início.1', 'Término.1', 'Alocação padrão']
                for _, row in df_filtrado.iterrows():
                    for j, coluna in enumerate(colunas, start=1):
                        valor = row.get(coluna, "")
                        # Verifica se o valor de "Andamento" é "Concluído" para posicionar "Alocação padrão" na coluna H
                        if coluna == 'Alocação padrão' and row.get('Andamento', '') == 'Concluído':
                            celula = nova_aba[f"H{ultima_linha}"]
                        else:
                            celula = nova_aba[f"{get_column_letter(j)}{ultima_linha}"]
                        celula.value = valor
                        logging.info(f"Escrevendo {valor} na célula {celula.coordinate}")
                    ultima_linha += 1

                # Inserir Totais
                total_linha = ultima_linha  # Linha onde será colocado o total
                bold_font = Font(bold=True, name="Tahoma", size=10)

                # Colocar "Total" na coluna F
                celula_total_texto = nova_aba[f"F{total_linha}"]
                celula_total_texto.value = "Total"
                celula_total_texto.font = bold_font

                # Calcular somas das colunas G e H
                celula_total_soma_g = nova_aba[f"G{total_linha}"]
                celula_total_soma_h = nova_aba[f"H{total_linha}"]
                celula_total_soma_g.value = f"=SUM(G9:G{ultima_linha - 1})"
                celula_total_soma_h.value = f"=SUM(H9:H{ultima_linha - 1})"

                # Definir o estilo das células de soma
                celula_total_soma_g.font = bold_font
                celula_total_soma_h.font = bold_font
                logging.info(f"Adicionado Total na linha {total_linha} para as colunas G e H.")

            progresso['value'] = i
            progresso.update_idletasks()

        pasta_dados = os.path.join(os.getcwd(), "Dados")
        if not os.path.exists(pasta_dados):
            os.makedirs(pasta_dados)
            logging.info(f"Pasta 'Dados' criada em {pasta_dados}")

        # Salvar o arquivo Excel com o nome contido no arquivo 'area.txt'
        nome_arquivo_saida = f"{area}.xlsx"
        novo_arquivo_excel = os.path.join(pasta_dados, nome_arquivo_saida)
        wb.save(novo_arquivo_excel)
        messagebox.showinfo("Sucesso", f"Planilha criada e salva em {novo_arquivo_excel}")
        logging.info(f"Arquivo Excel salvo em {novo_arquivo_excel}")

    except Exception as e:
        logging.error(f"Erro ao criar as abas no Excel: {e}")
        messagebox.showerror("Erro", "Ocorreu um erro ao criar as abas no Excel.")

# Função para criar a interface gráfica com barra de progresso
def criar_interface():
    app = ttk.Window(themename='cosmo')
    app.title("Sistema de Qualidade")
    app.geometry("600x400")
    app.resizable(True, True)

    # Título
    titulo = ttk.Label(app, text="Sistema de Qualidade", font=("Helvetica", 16), bootstyle=PRIMARY)
    titulo.pack(pady=10)

    # Botão para criar abas Excel
    btn_criar_abas = ttk.Button(app, text="Criar Abas no Excel", bootstyle=SUCCESS)
    btn_criar_abas.pack(pady=10)

    # Barra de progresso
    progresso = Progressbar(app, orient="horizontal", length=400, mode="determinate")
    progresso.pack(pady=20)

    # Ação ao clicar no botão "Criar Abas no Excel"
    btn_criar_abas.config(command=lambda: criar_abas_excel(progresso))

    app.mainloop()

if __name__ == "__main__":
    criar_interface()

